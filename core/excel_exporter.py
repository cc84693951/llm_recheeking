import os
import io
import tempfile
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.drawing.image import Image as XLImage


def export_results_to_excel(image_paths, get_results_func, save_path, img_size=(200, 150)):
    """
    将送检结果导出为 Excel 表格，包含渲染缩略图。

    :param image_paths: 图片路径列表
    :param get_results_func: 函数，接收 img_path，返回 (img_manager, bbox_manager) 或 (None, None)
    :param save_path: 保存的 xlsx 路径
    :param img_size: 插入 Excel 的图片缩略图尺寸
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "送检结果"

    # 表头
    headers = ["序号", "图片路径", "图片名称", "框序号", "Label", "X", "Y", "W", "H",
               "扩展后X", "扩展后Y", "扩展后W", "扩展后H", "复判结果", "渲染图"]
    ws.append(headers)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    row_idx = 2
    temp_files = []

    for img_idx, img_path in enumerate(image_paths, start=1):
        result = get_results_func(img_path)
        if not result or result[0] is None or result[1] is None:
            continue
        img_manager, bbox_manager, expand_settings = result

        # 绘制带结果的全图
        rendered = img_manager.draw_bboxes(bbox_manager, show_expanded=False, expand_settings={})
        if rendered is None:
            continue

        for bbox_idx, bbox in enumerate(bbox_manager.bboxes):
            # 计算扩展后坐标
            iw, ih = img_manager.size
            ex, ey, ew, eh = bbox.expanded_coords(
                iw, ih,
                center=expand_settings.get("center", 0),
                top=expand_settings.get("top", 1.0),
                bottom=expand_settings.get("bottom", 1.0),
                left=expand_settings.get("left", 1.0),
                right=expand_settings.get("right", 1.0),
            )

            ws.append([
                img_idx,
                img_path,
                os.path.basename(img_path),
                bbox_idx,
                bbox.label or "",
                bbox.x, bbox.y, bbox.w, bbox.h,
                ex, ey, ew, eh,
                bbox.result or "",
                ""
            ])

            # 为当前行所有单元格加边框
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).border = thin_border
                ws.cell(row=row_idx, column=col).alignment = Alignment(vertical="center", wrap_text=True)

            # 裁剪渲染图（仅当前框区域，带一点上下文）
            crop = rendered.crop((max(0, bbox.x - 10), max(0, bbox.y - 10),
                                   min(iw, bbox.right + 10), min(ih, bbox.bottom + 10)))
            crop = crop.resize(img_size, Image.LANCZOS)

            # 保存临时文件
            tmp_fd, tmp_path = tempfile.mkstemp(suffix=".png")
            os.close(tmp_fd)
            crop.save(tmp_path, format="PNG")
            temp_files.append(tmp_path)

            xl_img = XLImage(tmp_path)
            xl_img.width = img_size[0]
            xl_img.height = img_size[1]
            cell = ws.cell(row=row_idx, column=len(headers))
            ws.add_image(xl_img, cell.coordinate)

            # 设置行高以容纳图片
            ws.row_dimensions[row_idx].height = img_size[1] * 0.75 + 10

            row_idx += 1

    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # 图片列固定宽度
    ws.column_dimensions[ws.cell(row=1, column=len(headers)).column_letter].width = img_size[0] / 7 + 2

    wb.save(save_path)

    # 清理临时文件
    for tmp in temp_files:
        try:
            os.remove(tmp)
        except Exception:
            pass
